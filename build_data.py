# -*- coding: utf-8 -*-
"""
build_data.py — 从 Final_Match.xlsx + Precious_Metals_Summit_..._selection.xlsx
生成 dashboard/data.js 与 dashboard/companies_slug_map.md。

数据可重复生成：源 Excel 更新后直接重跑本脚本即可。
"""
import openpyxl, json, re, os, glob

BASE   = os.path.dirname(os.path.abspath(__file__))
FINAL  = os.path.join(BASE, "..", "Final_Match.xlsx")
SUMMIT = os.path.join(BASE, "..", "Precious_Metals_Summit_2026_Companies_selection.xlsx")
OUT_JS = os.path.join(BASE, "data.js")
OUT_MD = os.path.join(BASE, "companies_slug_map.md")

# ---- 17 种 primary commodity 配色 ----
COLORS = {
    "Gold":       "#e6b800",
    "Copper":     "#d35400",
    "Silver":     "#a7aeb6",
    "Nickel":     "#7f8c8d",
    "Zinc":       "#8e44ad",
    "Platinum":   "#2c3e50",
    "Palladium":  "#34495e",
    "Diamonds":   "#00bcd4",
    "Lithium":    "#16a085",
    "Cobalt":     "#2980b9",
    "Molybdenum": "#9b59b6",
    "Manganese":  "#e67e22",
    "Tantalum":   "#c0392b",
    "Vanadium":   "#27ae60",
    "Potash":     "#f39c12",
    "Bauxite":    "#a0522d",
    "Tin":        "#7f6c4a",
}

def slugify(name):
    s = str(name).lower().strip()
    s = re.sub(r"\.+$", "", s)          # 去尾点（规避 ATEX Resources Inc. 的尾点问题）
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")

def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def cell(v):
    return v if v is not None else ""

# ---- 1. Summit 公司（35 家，权威名单，生成 slug）----
wb = openpyxl.load_workbook(SUMMIT, data_only=True)
ws = wb.active
srows = list(ws.iter_rows(values_only=True))
companies = []
slug_map = {}
used_slug = set()
for r in srows[1:]:
    if r[0] is None or str(r[0]).strip() == "":
        continue
    name = str(r[0]).strip()
    slug = slugify(name)
    base = slug
    i = 2
    while slug in used_slug:
        slug = f"{base}-{i}"
        i += 1
    used_slug.add(slug)
    slug_map[name] = slug
    companies.append({
        "name": name,
        "website": cell(r[1]),
        "tickerRaw": cell(r[2]),
        "exchange": cell(r[3]),
        "tickerSymbols": cell(r[4]),
        "oneOnOne": cell(r[5]),
        "primaryFocus": cell(r[6]),
        "companyType": cell(r[7]),
        "commodity": cell(r[8]),
        "slug": slug,
        "mineCount": 0,
    })
wb.close()

# ---- 2. Final_Match 矿点（208 行）----
wb = openpyxl.load_workbook(FINAL, data_only=True)
ws = wb["Final_Match"]
frows = list(ws.iter_rows(values_only=True))
mines = []
for r in frows[1:]:
    if r[0] is None or str(r[0]).strip() == "":
        continue
    company_name = cell(r[38])
    slug = slug_map.get(company_name, slugify(company_name))
    mines.append({
        "propName": cell(r[0]),
        "commodity": cell(r[2]),
        "country": cell(r[6]),
        "stateProvince": cell(r[7]),
        "devStage": cell(r[8]),
        "actvStatus": cell(r[9]),
        "commoditiesList": cell(r[10]),
        "operator": cell(r[11]),
        "ownerName": cell(r[14]),
        "ownerType": cell(r[15]),
        "ownerCountry": cell(r[16]),
        "ownerList": cell(r[3]),
        "lat": num(r[30]),
        "lng": num(r[31]),
        "locationComments": cell(r[32]),
        "grdAuMI": num(r[20]),
        "containedAuMI": num(r[21]),
        "grdAuTotal": num(r[22]),
        "containedAuTotal": num(r[23]),
        "grdCuMI": num(r[24]),
        "containedCuMI": num(r[25]),
        "grdCuTotal": num(r[26]),
        "containedCuTotal": num(r[27]),
        "matchedOwnerString": cell(r[37]),
        "companyName": company_name,
        "exchange": cell(r[41]),
        "ticker": cell(r[42]),
        "website": cell(r[39]),
        "companySlug": slug,
    })
wb.close()

# 统计每家公司矿点数
name_to_company = {c["name"]: c for c in companies}
for m in mines:
    c = name_to_company.get(m["companyName"])
    if c:
        c["mineCount"] += 1

data = {
    "mines": mines,
    "companies": companies,
    "slugMap": slug_map,
    "colors": COLORS,
    # 自动扫描 companies/ 目录下已生成的公司 HTML，标识即可跳转（重跑脚本不会回退为空）
    "generatedCompanies": sorted(
        os.path.splitext(os.path.basename(f))[0]
        for f in glob.glob(os.path.join(BASE, "companies", "*.html"))
    ),
}

js = "window.DASHBOARD_DATA = " + json.dumps(data, ensure_ascii=False, indent=1) + ";\n"
with open(OUT_JS, "w", encoding="utf-8") as f:
    f.write(js)

# ---- 3. slug 映射清单（交给另一个 agent 对齐命名）----
lines = [
    "# 公司名 ↔ slug ↔ 链接 映射清单",
    "",
    "> 本文件由 `build_data.py` 自动生成，请勿手改。公司独立 html 请统一存到 `companies/<slug>.html`。",
    "",
    "| # | 公司名 | slug | 链接 | html 状态 |",
    "|---|--------|------|------|-----------|",
]
for i, c in enumerate(companies, 1):
    lines.append(f"| {i} | {c['name']} | `{c['slug']}` | `companies/{c['slug']}.html` | 生成中 |")
lines.append("")
with open(OUT_MD, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"mines: {len(mines)}")
print(f"companies: {len(companies)}")
with_mine = sum(1 for c in companies if c["mineCount"] > 0)
print(f"companies with >=1 mine: {with_mine}")
print(f"colors: {len(COLORS)}")
print(f"written: {OUT_JS}")
print(f"written: {OUT_MD}")
